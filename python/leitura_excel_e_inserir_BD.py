from openpyxl import load_workbook
import pyodbc

dados_conexao = ("driver=;""server=;""user=;""database=;""password=;")
cnx = pyodbc.connect(dados_conexao)
cursor = cnx.cursor()

wb = load_workbook(r'caminho do arquivo .xlsx')

planilha = wb["nome da planilha dentro do arquivo"]


for i in range (0, planilha.max_row):
    nome = planilha.cell(row=i+1, column=1).value
    idade = planilha.cell(row=i+1, column=2).value
    
    if nome and idade:
        comando = f"insert into infos (nome, idade) values ('{nome}', '{idade}')"
        cursor.execute(comando)
        cursor.commit()




'''
planilha.cell(column=1, row=1).value

planilha.max_row
planilha.max_column
'''